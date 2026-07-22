"""Implementación de PermitRepository sobre Excel (openpyxl).

Aplica el Repository Pattern: el caso de uso persiste permisos sin saber que el
almacén es un ``.xlsx``. Si el archivo no existe, lo crea con la fila de
encabezados. Cada permiso procesado se agrega como una fila nueva.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from permits.domain.entities.permit import Permit
from permits.domain.exceptions import PersistenceError
from permits.domain.ports.permit_repository import PermitRepository

logger = logging.getLogger(__name__)


class ExcelPermitRepository(PermitRepository):
    """Registra permisos procesados en un libro de Excel."""

    def __init__(self, *, excel_path: Path, sheet_name: str, headers: list[str]) -> None:
        self._path = Path(excel_path)
        self._sheet_name = sheet_name
        self._headers = headers

    def save(
        self,
        permit: Permit,
        *,
        processed_at: datetime,
        estado: str,
        observaciones: str = "",
    ) -> None:
        try:
            workbook, sheet = self._open_or_create()
            row = [
                processed_at.strftime("%Y-%m-%d %H:%M:%S"),
                str(permit.folio),
                permit.empresa,
                permit.date_range.start.isoformat(),
                permit.date_range.end.isoformat(),
                "Sí" if permit.vehiculo else "No",
                str(permit.plates),
                permit.cantidad_personas,
                estado,
                observaciones,
            ]
            sheet.append(row)
            self._path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(self._path)
            logger.info("Fila agregada a Excel: %s (%s)", permit.folio, self._path)
        except PersistenceError:
            raise
        except Exception as exc:
            raise PersistenceError(f"No se pudo escribir en Excel '{self._path}': {exc}") from exc

    # ------------------------------------------------------------------ #
    def _open_or_create(self) -> tuple[Workbook, object]:
        if self._path.is_file():
            workbook = load_workbook(self._path)
            if self._sheet_name in workbook.sheetnames:
                return workbook, workbook[self._sheet_name]
            sheet = workbook.create_sheet(self._sheet_name)
            sheet.append(self._headers)
            return workbook, sheet

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self._sheet_name
        sheet.append(self._headers)
        return workbook, sheet
