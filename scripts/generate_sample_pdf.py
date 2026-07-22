"""Genera los insumos del MVP en sample_data/.

Crea PDFs de solicitud de permiso (uno válido y uno inválido), un correo .eml de
ejemplo y un Excel maestro vacío con encabezados. Ejecutar:

    python scripts/generate_sample_pdf.py

Los PDFs generados tienen el formato de etiquetas que espera
``PdfplumberExtractor`` (``Etiqueta: valor`` + sección de personas con viñetas).
"""

from __future__ import annotations

import sys
from email.message import EmailMessage
from pathlib import Path

from fpdf import FPDF

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src"))

PDF_DIR = PROJECT_ROOT / "sample_data" / "pdfs"
EMAIL_DIR = PROJECT_ROOT / "sample_data" / "emails"
DATA_DIR = PROJECT_ROOT / "data"


def _write_permit_pdf(path: Path, fields: dict, personas: list[str]) -> None:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 12, "Solicitud de Permiso", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)
    pdf.set_font("Helvetica", "", 12)
    for label, value in fields.items():
        pdf.cell(0, 9, f"{label}: {value}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 9, "Personas con permiso:", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 12)
    for persona in personas:
        pdf.cell(0, 8, f"- {persona}", new_x="LMARGIN", new_y="NEXT")
    path.parent.mkdir(parents=True, exist_ok=True)
    pdf.output(str(path))
    print(f"[ok] PDF generado: {path.relative_to(PROJECT_ROOT)}")


def _write_sample_email(path: Path, pdf_name: str) -> None:
    msg = EmailMessage()
    msg["Subject"] = "Solicitud de Permiso"
    msg["From"] = "solicitante@empresa-cliente.com"
    msg["To"] = "permisos@miempresa.com"
    msg.set_content(
        "Buen dia,\n\nAdjunto la solicitud de permiso en PDF para su procesamiento.\n\n"
        "Saludos."
    )
    pdf_path = PDF_DIR / pdf_name
    msg.add_attachment(
        pdf_path.read_bytes(),
        maintype="application",
        subtype="pdf",
        filename=pdf_name,
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(bytes(msg))
    print(f"[ok] Email generado: {path.relative_to(PROJECT_ROOT)}")


def _write_empty_excel(path: Path) -> None:
    from openpyxl import Workbook

    from permits.config.settings import get_settings

    settings = get_settings()
    wb = Workbook()
    ws = wb.active
    ws.title = settings.excel.sheet_name
    ws.append(settings.excel.headers)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"[ok] Excel vacío generado: {path.relative_to(PROJECT_ROOT)}")


def main() -> None:
    # 1. Permiso VÁLIDO
    _write_permit_pdf(
        PDF_DIR / "permiso_ejemplo.pdf",
        {
            "Folio": "PER-001245",
            "Empresa": "Spin",
            "Fecha Inicio": "2026-01-10",
            "Fecha Fin": "2026-01-15",
            "Vehiculo": "Si",
            "Placas": "ABC123",
        },
        ["Juan Perez", "Maria Lopez"],
    )

    # 2. Permiso INVÁLIDO (vehículo = Sí pero sin placas)
    _write_permit_pdf(
        PDF_DIR / "permiso_invalido.pdf",
        {
            "Folio": "PER-009999",
            "Empresa": "ACME",
            "Fecha Inicio": "2026-02-20",
            "Fecha Fin": "2026-02-10",  # inicio > fin -> también inválido
            "Vehiculo": "Si",
            "Placas": "",
        },
        ["Pedro Ramirez"],
    )

    _write_sample_email(EMAIL_DIR / "correo_ejemplo.eml", "permiso_ejemplo.pdf")
    _write_empty_excel(DATA_DIR / "permisos.xlsx")
    print("\nInsumos del MVP listos en sample_data/ y data/.")


if __name__ == "__main__":
    main()
