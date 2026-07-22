"""Tests de integración de infraestructura: extractor real y repositorio Excel."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import pytest
from fpdf import FPDF
from openpyxl import load_workbook

from permits.domain.entities.permit import Permit
from permits.domain.exceptions import ExtractionError
from permits.infrastructure.extraction.pdfplumber_extractor import PdfplumberExtractor
from permits.infrastructure.persistence.excel_permit_repository import ExcelPermitRepository

HEADERS = ["Fecha proceso", "Folio", "Empresa", "Fecha Inicio", "Fecha Fin",
           "Vehículo", "Placas", "Cantidad Personas", "Estado", "Observaciones"]


def _make_permit_pdf(path: Path) -> None:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "", 12)
    for line in [
        "Solicitud de Permiso",
        "Folio: PER-777",
        "Empresa: Spin",
        "Fecha Inicio: 2026-01-10",
        "Fecha Fin: 2026-01-15",
        "Vehiculo: Si",
        "Placas: XYZ987",
        "Personas con permiso:",
        "- Ana Torres",
        "- Luis Diaz",
    ]:
        pdf.cell(0, 8, line, new_x="LMARGIN", new_y="NEXT")
    pdf.output(str(path))


class TestPdfplumberExtractor:
    def test_extracts_all_fields(self, tmp_path):
        pdf_path = tmp_path / "permit.pdf"
        _make_permit_pdf(pdf_path)

        dto = PdfplumberExtractor().extract(pdf_path)

        assert dto.folio == "PER-777"
        assert dto.empresa == "Spin"
        assert dto.fecha_inicio == date(2026, 1, 10)
        assert dto.fecha_fin == date(2026, 1, 15)
        assert dto.vehiculo is True
        assert dto.placas == "XYZ987"
        assert dto.personas == ["Ana Torres", "Luis Diaz"]

    def test_missing_file_raises(self, tmp_path):
        with pytest.raises(ExtractionError):
            PdfplumberExtractor().extract(tmp_path / "nope.pdf")


class TestExcelPermitRepository:
    def test_creates_file_with_headers_and_appends(self, tmp_path):
        excel_path = tmp_path / "out" / "permisos.xlsx"
        repo = ExcelPermitRepository(
            excel_path=excel_path, sheet_name="Permisos", headers=HEADERS
        )
        permit = Permit.create(
            folio="PER-777",
            empresa="Spin",
            fecha_inicio=date(2026, 1, 10),
            fecha_fin=date(2026, 1, 15),
            vehiculo=True,
            placas="XYZ987",
            personas=["Ana Torres", "Luis Diaz"],
        )
        repo.save(permit, processed_at=datetime(2026, 7, 21, 12, 0), estado="PROCESADO")

        assert excel_path.is_file()
        ws = load_workbook(excel_path).active
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == tuple(HEADERS)
        assert rows[1][1] == "PER-777"
        assert rows[1][7] == 2  # cantidad personas

    def test_appends_second_row(self, tmp_path):
        excel_path = tmp_path / "permisos.xlsx"
        repo = ExcelPermitRepository(
            excel_path=excel_path, sheet_name="Permisos", headers=HEADERS
        )
        permit = Permit.create(
            folio="PER-1", empresa="A", fecha_inicio=date(2026, 1, 1),
            fecha_fin=date(2026, 1, 2), vehiculo=False, placas="", personas=["X"],
        )
        repo.save(permit, processed_at=datetime(2026, 7, 21, 12, 0), estado="PROCESADO")
        repo.save(permit, processed_at=datetime(2026, 7, 21, 13, 0), estado="PROCESADO")

        rows = list(load_workbook(excel_path).active.iter_rows(values_only=True))
        assert len(rows) == 3  # encabezado + 2
